"""Generic spreadsheet serializer/parser for CSV, JSON, and XLSX.

All entity export/import handlers call these helpers. This is the only
module that imports openpyxl and csv — entity code stays format-agnostic.
"""
from __future__ import annotations

import csv
import io
import json
from typing import Any

from fastapi.responses import StreamingResponse

_IMPORT_ROW_LIMIT = 500
_IMPORT_FILE_SIZE_LIMIT = 5_000_000  # 5 MB


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------

def to_csv(rows: list[dict[str, Any]], columns: list[str]) -> bytes:
    """Serialize rows to UTF-8 CSV with BOM (Excel-compatible)."""
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=columns,
        extrasaction="ignore",
        lineterminator="\r\n",
    )
    writer.writeheader()
    for row in rows:
        writer.writerow({col: _cell(row.get(col)) for col in columns})
    # utf-8-sig BOM prefix so Excel opens without encoding dialog
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def to_json(rows: list[dict[str, Any]]) -> bytes:
    """Serialize rows to pretty-printed JSON."""
    return json.dumps(rows, indent=2, default=str).encode("utf-8")


def to_xlsx(
    rows: list[dict[str, Any]],
    columns: list[str],
    sheet_name: str = "Data",
) -> bytes:
    """Serialize rows to XLSX with frozen header row and auto column widths."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    # Write header
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=False)

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell(row.get(col_name)))

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-fit column widths (cap at 60)
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row in rows:
            val = str(_cell(row.get(col_name)) or "")
            max_len = max(max_len, min(len(val), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_xlsx_template(
    columns: list[str],
    examples: dict[str, Any],
    sheet_name: str = "Data",
    instructions: dict[str, str] | None = None,
) -> bytes:
    """Generate an annotated XLSX template with example row and instructions sheet."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    example_fill = PatternFill("solid", fgColor="F0F4F8")
    example_font = Font(italic=True, color="64748B", size=10)

    for col_idx, col_name in enumerate(columns, start=1):
        # Header
        hcell = ws.cell(row=1, column=col_idx, value=col_name)
        hcell.fill = header_fill
        hcell.font = header_font

        # Example row
        ecell = ws.cell(row=2, column=col_idx, value=examples.get(col_name, ""))
        ecell.fill = example_fill
        ecell.font = example_font

        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 4, 18)

    ws.freeze_panes = "A3"

    # Instructions sheet
    if instructions:
        ws2 = wb.create_sheet("How to Fill")
        ws2.cell(row=1, column=1, value="Column").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="Instructions").font = Font(bold=True)
        for row_idx, (col_name, desc) in enumerate(instructions.items(), start=2):
            ws2.cell(row=row_idx, column=1, value=col_name)
            ws2.cell(row=row_idx, column=2, value=desc)
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Import parser
# ---------------------------------------------------------------------------

def parse_import(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    """Parse CSV or JSON import file. Returns list of row dicts.

    Raises ValueError on: file too large, too many rows, unsupported format,
    or malformed content.
    """
    if len(file_bytes) > _IMPORT_FILE_SIZE_LIMIT:
        raise ValueError(
            f"File too large ({len(file_bytes):,} bytes). Maximum is {_IMPORT_FILE_SIZE_LIMIT:,} bytes."
        )

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        rows = _parse_csv(file_bytes)
    elif ext == "json":
        rows = _parse_json(file_bytes)
    else:
        raise ValueError(
            f"Unsupported file format '.{ext}'. Upload a .csv or .json file."
        )

    if len(rows) > _IMPORT_ROW_LIMIT:
        raise ValueError(
            f"Import limit is {_IMPORT_ROW_LIMIT} rows. File contains {len(rows)} rows. "
            "Split the file and import in batches."
        )

    return rows


def _parse_csv(data: bytes) -> list[dict[str, Any]]:
    # utf-8-sig handles BOM added by Excel
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    # Strip whitespace from keys and values
    return [
        {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
        for row in rows
    ]


def _parse_json(data: bytes) -> list[dict[str, Any]]:
    try:
        parsed = json.loads(data.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise ValueError(f"Invalid JSON: {exc}") from exc
    if not isinstance(parsed, list):
        raise ValueError("JSON import must be an array of objects at the top level.")
    return parsed


# ---------------------------------------------------------------------------
# Response builder
# ---------------------------------------------------------------------------

_MEDIA_TYPES = {
    "csv": "text/csv",
    "json": "application/json",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

_EXTENSIONS = {"csv": "csv", "json": "json", "xlsx": "xlsx"}


def make_streaming_response(
    data: bytes,
    fmt: str,
    filename_stem: str,
) -> StreamingResponse:
    """Build a StreamingResponse with correct Content-Type and Content-Disposition."""
    media_type = _MEDIA_TYPES.get(fmt, "application/octet-stream")
    ext = _EXTENSIONS.get(fmt, fmt)
    filename = f"{filename_stem}.{ext}"
    return StreamingResponse(
        iter([data]),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell(value: Any) -> Any:
    """Normalize a cell value for CSV/XLSX output."""
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value)
    return value
